import pandas as pd
import os
import re
import sys
import json
from tqdm import tqdm
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import rdflib
from rdflib import Graph, Literal, Namespace
from jinja2 import Template, Environment, FileSystemLoader, StrictUndefined, Undefined
import importlib
import requests

load_dotenv()
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(parent_dir)


def load_config(config_path="config.json"):
    """設定ファイルを読み込む"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)



class WarnOnUndefined(Undefined):
    def _fail_with_undefined_error(self, *args, **kwargs):
        print(f"[警告] 未定義の変数にアクセスしました: {self._undefined_name}")
        return f"<<{self._undefined_name}>>"  # プレースホルダとしても見えるように

    __add__ = __radd__ = __mul__ = __rmul__ = __str__ = _fail_with_undefined_error
    __getitem__ = __getattr__ = _fail_with_undefined_error


EPSILON = 0.1

def init_ollama_llm(base_url="http://localhost:11434"):
    return {"base_url": base_url}

def detect_ontology_format(file_path):
    """Detect if the ontology file is JSON-LD"""
    _, ext = os.path.splitext(file_path)
    if ext == ".jsonld":
        return 'jsonld'
    elif ext == '.nl':
        return 'nl'
    else:
        raise ValueError(f"Unsupported ontology format in {file_path}: {ext}")

def read_ontology(file_path):
    """Read ontology file in either JSON-LD"""
    format_type = detect_ontology_format(file_path)
    
    if format_type == 'jsonld':
        with open(file_path, 'r') as f:
            return json.load(f)
    elif format_type == 'nl':
        with open(file_path, 'r') as f:
            return f.read()
        

def render_nl(template_path, comments_path):
    with open(template_path, 'r', encoding='utf-8') as f:
        template_str = f.read()

    with open(comments_path, 'r', encoding='utf-8') as f:
        comments = json.load(f)

    env = Environment(undefined=WarnOnUndefined)
    template = env.from_string(template_str)
    rendered = template.render(comments=comments)
    return rendered


def render_jsonld(template_path, comments_path):
    with open(template_path, 'r', encoding='utf-8') as f:
        template_str = f.read()

    with open(comments_path, 'r', encoding='utf-8') as f:
        comments = json.load(f)

    env = Environment(undefined=WarnOnUndefined)
    template = env.from_string(template_str)
    rendered = template.render(comments=comments)

    parsed = json.loads(rendered)
    return json.dumps(parsed, ensure_ascii=False, indent=2)

def get_pdf_text(df, rows_of_interest, file_name, text_type):
    file_df = df[df['file_name'] == file_name]
    label_data_dict = {}
    concatenated_text = ""
    if not file_df.empty:
        row = file_df.iloc[0]
        for col in rows_of_interest:
            if col in row:
                label_data_dict[col] = row[col]
        text_columns = [col for col in file_df.columns if col.startswith(f"text ({text_type})")]
        text_columns = sorted(text_columns, key=lambda x: int(x.split("page")[-1]))
        text_pieces = []
        for col in text_columns:
            value = row[col]
            if pd.notna(value):
                text_pieces.append(str(value))
        concatenated_text = "\n\n".join(text_pieces)
    return label_data_dict, concatenated_text

def get_ollama_response_for_pdf_text(client, pdf_text, ontology_template, comment_file, ontology_format='jsonld', temperature=0.1, model="gemma3n:e4b"):
    if len(pdf_text) < 10:
        print(pdf_text)
    
    with open(ontology_example_path, "r") as f:
        example_text = f.read()
    
    if ontology_format == 'jsonld':
        with open(comment_file, 'r', encoding='utf-8') as f:
            comments = json.load(f)
        ontology_text = render_jsonld(ontology_template, comment_file)
        docuemnt_info = comments["ut-trust:docInfo"]
        system_prompt = f"Extract information from the following document according to the JSON-LD below, and output it in JSON format. Note that not all entities would be present. Note that not all items would be present. Please take into account the following document detail when extracting the information: {docuemnt_info}"
        user_prompt = "\"""JSON-LD" + ontology_text + "\n\"""" + "\n\nExample output format:\n\"""json\n" + example_text + "\n\"""\n\nDocument to extract from:\n" + pdf_text
    elif ontology_format == 'nl':
        with open(comment_file, 'r', encoding='utf-8') as f:
            comments = json.load(f)
        ontology_text = render_nl(ontology_template, comment_file)
        docuemnt_info = comments["ut-trust:docInfo"]
        system_prompt = f"Extract information from the following Text. Extract items are listed below. Output must be in key-value single level JSON format. Note that not all items would be present. Please take into account the following document detail when extracting the information: {docuemnt_info}"
        user_prompt = "\"""Items to be extracted\n" + ontology_text + "\n\nExample output format:\n\"""json\n" + example_text + "\nText：\n" + pdf_text
    
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt}
    ]
    
    payload = {
        "model": model,
        "messages": messages,
        "stream": False,
        "options": {
            "temperature": temperature
        }
    }
    
    response = requests.post(f"{client['base_url']}/api/chat", json=payload)
    response.raise_for_status()
    
    response_data = response.json()
    output_content = response_data["message"]["content"]

    input_prompt = system_prompt + "\n\n" + user_prompt
    
    return input_prompt, output_content


def export_llm_log(llm_input, llm_output, out_path):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("=== llm_input start ===\n")
        f.write(llm_input)
        f.write("\n=== llm_input end ===\n")
        f.write("\n=== llm_output start ===\n")
        f.write(llm_output)
        f.write("\n=== llm_output end ===")


def get_dict_from_llm_response(rule_module, llm_output):
    try:
        cleaned = llm_output.strip().removeprefix("```json").removesuffix("```").strip()
        llm_output_dict = json.loads(cleaned)
    except json.JSONDecodeError:
        print("Error decoding JSON from LLM output.")
        print("\n-----")
        print(llm_output)
        print("-----\n")
        return {}
    
    converter = importlib.import_module(rule_module)
    convert = getattr(converter, "convert")
    
    try:
        row = convert(llm_output_dict)
        return row
    except Exception as e:
        print("Error happened while get_dict_from_llm_response", e)
        return None

def date_conversion(in_str):
    if str(in_str) == "No Information":
        return in_str
    return re.sub(r'\D', '', str(in_str))


def number_conversion(in_num):
    if str(in_num) == "No Information":
        return in_num
    elif isinstance(in_num, int) or isinstance(in_num, float):
       return abs(in_num)
    try:
        return abs(float(in_num))
    except:
        return "No Information (type error)"


def apply_rule_func(func, arg):
    try:
        return func(arg)
    except Exception as e:
        # print(f"Exception during apply_rule_func: {e}")
        return True


def append_dict_to_excel(file_name, headers, excel_path, llm_dict, labels_dict):
    combined_data = {"file_name": file_name}
    for header in headers:
        llm_column_name = f"llm_{header}"
        correct_column_name = f"correct_{header}"
        combined_data[llm_column_name] = llm_dict.get(header, "")
        combined_data[correct_column_name] = labels_dict.get(header, "")
    
    combined_df = pd.DataFrame([combined_data])
    df_color = make_color_df(combined_df)
    
    if not excel_path.exists():
        llm_headers = [f"llm_{header}" for header in headers]
        correct_headers = [f"correct_{header}" for header in headers]
        all_headers = [None]*(2*len(headers)+1)
        all_headers[0] = "file_name"
        all_headers[1::2] = llm_headers
        all_headers[2::2] = correct_headers
        
        df_header = pd.DataFrame(columns=all_headers)
        df_header.to_excel(excel_path, index=False)
    
    with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        existing_df = pd.read_excel(excel_path)
        startrow = existing_df.shape[0] + 1
        combined_df.to_excel(writer, index=False, header=False, startrow=startrow)

    wb = load_workbook(excel_path)
    ws = wb.active
    for row_idx, row in enumerate(df_color.itertuples(index=False), start=1):  # Excelは1始まり、1行目はヘッダ
        for col_idx, color in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + startrow, column=col_idx)
            cell.font = Font(color=color)
    wb.save(excel_path)

def make_color_df(df):
    # 同じshapeのdf_colorを作成（初期値は黒）
    df_color = pd.DataFrame("000000", index=df.index, columns=df.columns)
    
    # 奇数列とその右隣の列を比較して色を決定
    columns = df.columns
    for i in range(1, len(columns), 2):  # 1, 3, 5, ... が奇数インデックス扱い
        if i >= len(columns):
            break
            
        col_left = columns[i]
        col_right = columns[i+1]
        
        # セル単位で比較して色を設定
        for idx in df.index:
            if not is_equal(df.at[idx, col_left], df.at[idx, col_right]):
                df_color.at[idx, col_left] = "FF0000"    
    return df_color
    
def is_equal(val1, val2):
    if pd.isna(val1) and pd.isna(val2):
        return True
    try:
        return abs(float(val1) - float(val2)) < 1e-5
    except:
        return val1 == val2


if __name__ == "__main__":
    # 設定ファイルを読み込み
    config = load_config()
    
    ####### CONFIG FROM FILE ######## 
    N_EXP = config["experiment"]["n_exp"]
    temperature = config["experiment"]["temperature"]
    ontology_type = config["ontology"]["type"]
    ontology_file_expression = config["ontology"]["file_expression"]
    doc_type = config["document"]["type"]
    dict_version = config["document"]["dict_version"]
    label_with_data_path = config["paths"]["label_with_data_path"]
    prompt_dir = config["paths"]["prompt_dir"]
    output_base_path = Path(config["paths"]["output_base_path"])
    OLLAMA_MODEL = config["llm"]["ollama"]["model"]
    OLLAMA_URL = config["llm"]["ollama"]["url"]
    
    ####### CONSTANT FROM CONFIG ######## 
    ontology_type_config = config["ontology_type_config"]
    doc_type_config = config["doc_type_config"]
    
    #########################
    client = init_ollama_llm(OLLAMA_URL)
    ####### REQUIRED SETTINGS ########
    doc_abbrebiation = doc_type_config[doc_type]["abbrebiation"]
    dict_base_path = Path(config["paths"]["dict_base_path"])
    ontology_base_path = Path(config["paths"]["ontology_base_path"])
    ontology_example_base_path = Path(config["paths"]["ontology_example_base_path"])
    rule_base = "template.rule"
    
    ontology_file_pattern = re.compile(ontology_file_expression)
    ontology_matching_files = [ontology_base_path / ontology_type / filename for filename in os.listdir(str(ontology_base_path/ontology_type)) if ontology_file_pattern.match(filename)]
    required_headers = doc_type_config[doc_type]["required_headers"]
    file_ids = doc_type_config[doc_type]["num_ids"]
    dict_path = dict_base_path / f"{dict_version}.json"
    
    print(f"ontology_type: {ontology_type} doc_type: {doc_type}")
    ##################################
    for num in range(N_EXP):
        print(f"----Trial : {num+1} ----")
        for ontology_definition_path in ontology_matching_files:
            print(f"ontology file: {ontology_definition_path}")
            ontology_file_without_ext = Path(ontology_definition_path).stem
            ontology_example_path = ontology_example_base_path / ontology_type / f"{ontology_file_without_ext}.txt"
            rule_module = rule_base + f".{ontology_type}" + f".{ontology_file_without_ext}"
            output_excel_path = output_base_path / ontology_type / f"{doc_abbrebiation}_{ontology_type}_{ontology_file_without_ext}_dict_{dict_version}.xlsx"
            if not output_excel_path.parent.exists():
                output_excel_path.parent.mkdir(parents=True, exist_ok=True)

            df = pd.read_excel(label_with_data_path, engine="openpyxl", sheet_name=doc_type)
            file_names = [f"{i}.pdf" for i in file_ids]
            for file_name in tqdm(file_names, desc="Processing PDFs"):
                labels_dict, pdf_text = get_pdf_text(df, required_headers, file_name, "ocr")
                
                llm_input, llm_output = get_ollama_response_for_pdf_text(client, pdf_text, ontology_definition_path, dict_path, ontology_type, temperature, OLLAMA_MODEL)
                
                if True:
                    out_path = Path(prompt_dir) / Path(doc_type) / Path(ontology_type) / Path(f"{os.path.splitext(file_name)[0]}_{ontology_file_without_ext}_{num+1}.txt")
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    export_llm_log(llm_input,llm_output,out_path)
                llm_dict = get_dict_from_llm_response(rule_module, llm_output)
                if not llm_dict:
                    print(f"Skip processing due to error: {file_name}")
                    continue

                append_dict_to_excel(file_name, required_headers, output_excel_path, llm_dict, labels_dict)